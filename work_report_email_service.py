import os
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pymysql
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
# We explicitly load the base .env first, then APP_ENV file, to match the main.py behavior
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Database credentials
DB_HOST = "localhost"
DB_USER = "admin"
DB_PASSWORD = "01eMatrix007!" # Please change according to your local DB
DB_NAME = "tms"

# SMTP and Mail Configuration
SMTP_SERVER="s11777.bom1.stableserver.net"
SMTP_PORT=587
SENDER_EMAIL="supportdesk@ematrixinfotechpms.com"
FROM_EMAIL="supportdesk@ematrixinfotechpms.com"
REPLY_TO="supportdesk@ematrixinfotechpms.com"
SENDER_PASSWORD="01eMatrix007!"
SEND_MAIL_TO="ritesh@ematrixinfotech.com"

def get_db_connection():
    return pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

def format_time_display(hours, minutes):
    h = int(hours or 0)
    m = int(minutes or 0)
    h_str = f"{h} hour" if h == 1 else f"{h} hours"
    m_str = f"{m} minute" if m == 1 else f"{m} minutes"
    if h > 0 and m > 0:
        return f"{h_str} {m_str}"
    elif h > 0:
        return h_str
    else:
        return m_str

def generate_report_excel(records, date_str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Work Report"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Define design colors and styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FF172B4D")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True, color="FF172B4D")
    
    header_fill = PatternFill(start_color="FF0052CC", end_color="FF0052CC", fill_type="solid") # Jira Blue
    
    thin_border = Border(
        left=Side(style='thin', color='FF808080'),
        right=Side(style='thin', color='FF808080'),
        top=Side(style='thin', color='FF808080'),
        bottom=Side(style='thin', color='FF808080')
    )
    
    # Title Row (A1:H1)
    ws.merge_cells("A1:H1")
    for col_idx in range(1, 9):
        ws.cell(row=1, column=col_idx).border = thin_border
    title_cell = ws["A1"]
    title_cell.value = f"Daily Work Report - {date_str}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Blank Row (A2:H2)
    for col_idx in range(1, 9):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = ""
        cell.border = thin_border
    ws.row_dimensions[2].height = 20
    
    # Headers (8 columns including Note)
    headers = [
        "Developer Name", "Ticket No", "Project Name", "Ticket Title", 
        "Ticket Status", "Worked Time", "Actual Time", "Note"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 25
    
    # Group records by Developer Name
    dev_map = {}
    for r in records:
        dev_name = f"{r.get('first_name', '')} {r.get('last_name', '')}".strip()
        if dev_name not in dev_map:
            dev_map[dev_name] = []
        dev_map[dev_name].append(r)
        
    # Write Data
    row_idx = 4
    for dev_name, tickets in dev_map.items():
        start_row = row_idx
        num_tickets = len(tickets)
        end_row = start_row + num_tickets - 1
        
        for i, r in enumerate(tickets):
            r_row = start_row + i
            
            # Developer Name cell
            dev_cell = ws.cell(row=r_row, column=1)
            dev_cell.value = dev_name
            dev_cell.font = bold_font
            dev_cell.border = thin_border
            dev_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Times formatting
            tw_mins = int(r.get("total_worked_minutes") or 0)
            worked_time_display = format_time_display(tw_mins // 60, tw_mins % 60)
            
            act_secs = int(r.get("total_actual_seconds") or 0)
            actual_time_display = format_time_display(act_secs // 3600, (act_secs % 3600) // 60)
            
            row_data = [
                (r.get("ticket_no", ""), bold_font, Alignment(horizontal="left", vertical="center")),
                (r.get("project_name", ""), bold_font, Alignment(horizontal="left", vertical="center")),
                (r.get("ticket_title", ""), bold_font, Alignment(horizontal="left", vertical="center")),
                (r.get("ticket_status", ""), bold_font, Alignment(horizontal="center", vertical="center")),
                (worked_time_display, bold_font, Alignment(horizontal="center", vertical="center")),
                (actual_time_display, bold_font, Alignment(horizontal="center", vertical="center")),
                (r.get("note", "") or "", bold_font, Alignment(horizontal="left", vertical="center")),
            ]
            
            for col_offset, (val, fnt, align) in enumerate(row_data, start=2):
                cell = ws.cell(row=r_row, column=col_offset)
                cell.value = val
                cell.font = fnt
                cell.border = thin_border
                cell.alignment = align
                
            ws.row_dimensions[r_row].height = 24
            
        if num_tickets > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            
        row_idx = end_row + 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    file_path = os.path.join(BASE_DIR, f"Daily_Work_Report_{date_str}.xlsx")
    wb.save(file_path)
    return file_path

def send_report_email(file_path, date_str):
    if not SEND_MAIL_TO:
        raise ValueError("SEND_MAIL_TO is not set in environment variables")
        
    msg = MIMEMultipart()
    msg['From'] = f"DeskEmatrixInfoTech <{SENDER_EMAIL}>"
    msg['To'] = SEND_MAIL_TO
    msg['Subject'] = f"Daily Work Report - {date_str}"
    if REPLY_TO:
        msg['Reply-To'] = REPLY_TO
        
    body = f"Please find attached the daily work report for {date_str}."
    msg.attach(MIMEText(body, 'plain'))
    # Attach Excel file
    with open(file_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {os.path.basename(file_path)}",
        )
        msg.attach(part)
        
    try:
        print("Connecting to SMTP server...")
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        # server.set_debuglevel(1)  # Enable detailed SMTP debug output
        
        print("Starting TLS...")
        server.starttls()
        
        print(f"Logging in as {SENDER_EMAIL}...")
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        
        print(f"Sending email from {SENDER_EMAIL} to {SEND_MAIL_TO}...")
        refused = server.sendmail(SENDER_EMAIL, SEND_MAIL_TO, msg.as_string())
        if refused:
            print(f"Mail delivery refused for some recipients: {refused}")
            
        server.quit()
        print(f"Report email successfully sent to {SEND_MAIL_TO}")
    except Exception as e:
        print(f"Failed to send email: {e}")

def run_daily_report_service(target_date=None):
    if target_date is None:
        target_date = datetime.date.today()
    elif isinstance(target_date, str):
        target_date = datetime.datetime.strptime(target_date, "%Y-%m-%d").date()
        
    date_str = target_date.strftime("%Y-%m-%d")
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # Query grouped unique developer-ticket metrics for target_date
            sql = """
                SELECT 
                    ut.user_id,
                    ut.ticket_id,
                    u.first_name,
                    u.last_name,
                    u.email,
                    t.project_id,
                    p.name AS project_name,
                    t.title AS ticket_title,
                    t.ticket_no,
                    st.name AS ticket_status,
                    COALESCE(ttw_summary.total_worked_minutes, 0) AS total_worked_minutes,
                    COALESCE(ttl_summary.total_actual_seconds, 0) AS total_actual_seconds,
                    COALESCE(ttw_summary.note, '') AS note
                FROM (
                    SELECT user_id, ticket_id FROM today_ticket_work WHERE date = %s
                    UNION
                    SELECT user_id, ticket_id FROM ticket_time_log WHERE DATE(start_time) = %s AND end_time IS NOT NULL
                ) ut
                JOIN users u ON ut.user_id = u.id
                JOIN tickets t ON ut.ticket_id = t.id
                LEFT JOIN projects p ON t.project_id = p.id
                LEFT JOIN status st ON t.status_id = st.id
                LEFT JOIN (
                    SELECT user_id, ticket_id, 
                           SUM(CAST(COALESCE(hours, 0) AS UNSIGNED) * 60 + CAST(COALESCE(minutes, 0) AS UNSIGNED)) AS total_worked_minutes,
                           GROUP_CONCAT(DISTINCT note SEPARATOR '; ') AS note
                    FROM today_ticket_work
                    WHERE date = %s
                    GROUP BY user_id, ticket_id
                ) ttw_summary ON ut.user_id = ttw_summary.user_id AND ut.ticket_id = ttw_summary.ticket_id
                LEFT JOIN (
                    SELECT user_id, ticket_id, SUM(TIMESTAMPDIFF(SECOND, start_time, end_time)) AS total_actual_seconds
                    FROM ticket_time_log
                    WHERE DATE(start_time) = %s AND end_time IS NOT NULL AND status != 1
                    GROUP BY user_id, ticket_id
                ) ttl_summary ON ut.user_id = ttl_summary.user_id AND ut.ticket_id = ttl_summary.ticket_id
                ORDER BY u.id ASC, t.ticket_no ASC
            """
            cursor.execute(sql, (date_str, date_str, date_str, date_str))
            records = cursor.fetchall()
            if not records:
                print(f"No records found for date {date_str}. Email not sent.")
                return False
                
            # Generate Excel
            file_path = generate_report_excel(records, date_str)
            
            # Send Email
            send_report_email(file_path, date_str)
            
            # Clean up generated Excel file
            if os.path.exists(file_path):
                os.remove(file_path)
                
            return True
    finally:
        conn.close()

if __name__ == "__main__":
    # If run directly, run for today
    run_daily_report_service()
