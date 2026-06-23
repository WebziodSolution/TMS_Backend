from fastapi import HTTPException
from database import get_db_connection

class ReportsService:
    @staticmethod
    def get_daily_report(date: str, login_user_id: int):
        conn = get_db_connection()
        try:
            with conn.cursor() as cursor:
                # Check user role
                cursor.execute("SELECT role_id FROM users WHERE id = %s", (login_user_id,))
                user_role_rec = cursor.fetchone()
                if not user_role_rec:
                    raise HTTPException(status_code=404, detail="User not found")
                role_id = user_role_rec['role_id']

                role_filter = ""
                params = []
                if role_id == 2: # Developer
                    role_filter = "AND tw.user_id = %s"
                    params = [date, login_user_id]
                elif role_id == 4: # Manager
                    role_filter = "AND (tw.user_id = %s OR u.report_to = %s)"
                    params = [date, login_user_id, login_user_id]
                else: # Admin / Administrator / Others
                    role_filter = ""
                    params = [date]

                # Query daily tickets worked on by developers for the given date
                sql = f"""
                    SELECT 
                        tw.user_id as developer_id,
                        CONCAT(u.first_name, ' ', u.last_name) as developer_name,
                        u.work_hours as total_working_hours,
                        tw.ticket_id,
                        t.ticket_no,
                        p.name as project_name,
                        t.title as ticket_title,
                        tw.hours,
                        tw.minutes,
                        (
                            SELECT COALESCE(SUM(TIMESTAMPDIFF(SECOND, tl.start_time, tl.end_time)), 0)
                            FROM ticket_log tl
                            WHERE tl.ticket_id = tw.ticket_id 
                            AND tl.status != 1
                            AND tl.user_id = tw.user_id 
                            AND DATE(tl.start_time) = tw.date
                        ) as actual_work_seconds
                    FROM today_ticket_work tw
                    JOIN users u ON tw.user_id = u.id
                    JOIN tickets t ON tw.ticket_id = t.id
                    LEFT JOIN projects p ON t.project_id = p.id
                    WHERE tw.date = %s {role_filter}
                    ORDER BY u.id ASC, t.ticket_no ASC
                """
                cursor.execute(sql, tuple(params))
                results = cursor.fetchall()
                
                developers = {}
                for row in results:
                    dev_id = row['developer_id']
                    if dev_id not in developers:
                        developers[dev_id] = {
                            "developer_id": str(dev_id),
                            "developer_name": row['developer_name'],
                            "total_working_hours": str(row['total_working_hours']) if row['total_working_hours'] is not None else "0.0",
                            "tickets": []
                        }
                    
                    # Formatting hours and minutes to something like 1.20 for 1 hour 20 minutes
                    hours_val = int(row['hours'] or 0)
                    minutes_val = int(row['minutes'] or 0)
                    today_working_hours = f"{hours_val}.{minutes_val:02d}"
                    
                    # Format actual work seconds as "H.MM"
                    actual_seconds = int(row.get('actual_work_seconds') or 0)
                    actual_h = actual_seconds // 3600
                    actual_m = (actual_seconds % 3600) // 60
                    actual_working_hours = f"{actual_h}.{actual_m:02d}"
                    
                    developers[dev_id]["tickets"].append({
                        "ticket_id": str(row['ticket_id']),
                        "ticket_no": row['ticket_no'],
                        "project_name": row['project_name'],
                        "title": row['ticket_title'],
                        "today_working_hours": today_working_hours,
                        "actual_working_hours": actual_working_hours
                    })
                
                return list(developers.values())
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        finally:
            conn.close()

    @staticmethod
    def get_monthly_report(start_date: str, end_date: str, group_by: str = "ticket"):
        conn = get_db_connection()
        try:
            with conn.cursor() as cursor:
                # Query daily tickets worked on by developers between start_date and end_date
                sql = """
                    SELECT 
                        tw.ticket_id,
                        t.ticket_no,
                        p.name as project_name,
                        t.title as ticket_title,
                        tw.user_id as developer_id,
                        CONCAT(u.first_name, ' ', u.last_name) as developer_name,
                        tw.hours,
                        tw.minutes,
                        (
                            SELECT COALESCE(SUM(TIMESTAMPDIFF(SECOND, tl.start_time, tl.end_time)), 0)
                            FROM ticket_log tl
                            WHERE tl.ticket_id = tw.ticket_id 
                            AND tl.status != 1
                            AND tl.user_id = tw.user_id 
                            AND DATE(tl.start_time) = tw.date
                        ) as actual_work_seconds
                    FROM today_ticket_work tw
                    JOIN users u ON tw.user_id = u.id
                    JOIN tickets t ON tw.ticket_id = t.id
                    LEFT JOIN projects p ON t.project_id = p.id
                    WHERE tw.date BETWEEN %s AND %s
                    ORDER BY t.ticket_no ASC, u.id ASC
                """
                cursor.execute(sql, (start_date, end_date))
                results = cursor.fetchall()
                
                if group_by == "developer":
                    # Group by developer, and sum hours for each developer and each ticket for that developer
                    developers = {}
                    for row in results:
                        dev_id = row['developer_id']
                        if dev_id not in developers:
                            developers[dev_id] = {
                                "developer_id": str(dev_id),
                                "developer_name": row['developer_name'],
                                "total_minutes": 0,
                                "actual_seconds": 0,
                                "tickets_map": {}
                            }
                        
                        # Convert to minutes for precise summation
                        hours_val = int(row['hours'] or 0)
                        minutes_val = int(row['minutes'] or 0)
                        row_mins = hours_val * 60 + minutes_val
                        
                        developers[dev_id]["total_minutes"] += row_mins
                        developers[dev_id]["actual_seconds"] += int(row['actual_work_seconds'] or 0)
                        
                        t_id = row['ticket_id']
                        if t_id not in developers[dev_id]["tickets_map"]:
                            developers[dev_id]["tickets_map"][t_id] = {
                                "ticket_id": str(t_id),
                                "ticket_no": row['ticket_no'],
                                "project_name": row['project_name'],
                                "title": row['ticket_title'],
                                "minutes": 0,
                                "actual_seconds": 0
                            }
                        developers[dev_id]["tickets_map"][t_id]["minutes"] += row_mins
                        developers[dev_id]["tickets_map"][t_id]["actual_seconds"] += int(row['actual_work_seconds'] or 0)
                    
                    # Sort developers by developer_name
                    sorted_dev_ids = sorted(developers.keys(), key=lambda d: developers[d]["developer_name"].lower())
                    
                    # Build the response list
                    report = []
                    for dev_id in sorted_dev_ids:
                        dev_data = developers[dev_id]
                        total_h = dev_data["total_minutes"] // 60
                        total_m = dev_data["total_minutes"] % 60
                        total_working_hours = f"{total_h}.{total_m:02d}"
                        
                        total_actual_h = dev_data["actual_seconds"] // 3600
                        total_actual_m = (dev_data["actual_seconds"] % 3600) // 60
                        total_actual_hours = f"{total_actual_h}.{total_actual_m:02d}"
                        
                        # Build tickets list
                        tickets_list = []
                        # Sort tickets by ticket_no ASC
                        sorted_t_ids = sorted(dev_data["tickets_map"].keys(), key=lambda tid: dev_data["tickets_map"][tid]["ticket_no"])
                        for t_id in sorted_t_ids:
                            t_data = dev_data["tickets_map"][t_id]
                            th = t_data["minutes"] // 60
                            tm = t_data["minutes"] % 60
                            working_hours = f"{th}.{tm:02d}"
                            
                            ta_h = t_data["actual_seconds"] // 3600
                            ta_m = (t_data["actual_seconds"] % 3600) // 60
                            actual_working_hours = f"{ta_h}.{ta_m:02d}"
                            
                            tickets_list.append({
                                "ticket_id": t_data["ticket_id"],
                                "ticket_no": t_data["ticket_no"],
                                "project_name": t_data["project_name"],
                                "title": t_data["title"],
                                "working_hours": working_hours,
                                "actual_working_hours": actual_working_hours
                            })
                        
                        report.append({
                            "developer_id": dev_data["developer_id"],
                            "developer_name": dev_data["developer_name"],
                            "total_working_hours": total_working_hours,
                            "total_actual_hours": total_actual_hours,
                            "tickets": tickets_list
                        })
                    
                    return report

                else:
                    # Group by ticket, and sum hours for each ticket and each user on that ticket
                    tickets = {}
                    for row in results:
                        t_id = row['ticket_id']
                        if t_id not in tickets:
                            tickets[t_id] = {
                                "ticket_id": str(t_id),
                                "ticket_no": row['ticket_no'],
                                "project_name": row['project_name'],
                                "title": row['ticket_title'],
                                "total_minutes": 0,
                                "actual_seconds": 0,
                                "users_map": {}
                            }
                        
                        # Convert to minutes for precise summation
                        hours_val = int(row['hours'] or 0)
                        minutes_val = int(row['minutes'] or 0)
                        row_mins = hours_val * 60 + minutes_val
                        
                        tickets[t_id]["total_minutes"] += row_mins
                        tickets[t_id]["actual_seconds"] += int(row['actual_work_seconds'] or 0)
                        
                        dev_id = row['developer_id']
                        if dev_id not in tickets[t_id]["users_map"]:
                            tickets[t_id]["users_map"][dev_id] = {
                                "developer_id": str(dev_id),
                                "developer_name": row['developer_name'],
                                "minutes": 0,
                                "actual_seconds": 0
                            }
                        tickets[t_id]["users_map"][dev_id]["minutes"] += row_mins
                        tickets[t_id]["users_map"][dev_id]["actual_seconds"] += int(row['actual_work_seconds'] or 0)
                    
                    # Build the response list
                    report = []
                    for t_id, t_data in tickets.items():
                        # Format total working hours
                        total_h = t_data["total_minutes"] // 60
                        total_m = t_data["total_minutes"] % 60
                        total_working_hours = f"{total_h}.{total_m:02d}"
                        
                        # Format total actual working hours
                        total_actual_h = t_data["actual_seconds"] // 3600
                        total_actual_m = (t_data["actual_seconds"] % 3600) // 60
                        total_actual_hours = f"{total_actual_h}.{total_actual_m:02d}"
                        
                        # Build users list
                        users_list = []
                        for dev_id, u_data in t_data["users_map"].items():
                            uh = u_data["minutes"] // 60
                            um = u_data["minutes"] % 60
                            working_hours = f"{uh}.{um:02d}"
                            
                            ua_h = u_data["actual_seconds"] // 3600
                            ua_m = (u_data["actual_seconds"] % 3600) // 60
                            actual_working_hours = f"{ua_h}.{ua_m:02d}"
                            
                            users_list.append({
                                "developer_id": u_data["developer_id"],
                                "developer_name": u_data["developer_name"],
                                "working_hours": working_hours,
                                "actual_working_hours": actual_working_hours
                            })
                        
                        report.append({
                            "ticket_id": t_data["ticket_id"],
                            "ticket_no": t_data["ticket_no"],
                            "project_name": t_data["project_name"],
                            "title": t_data["title"],
                            "total_working_hours": total_working_hours,
                            "total_actual_hours": total_actual_hours,
                            "users": users_list
                        })
                    
                    return report
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        finally:
            conn.close()

    @staticmethod
    def export_monthly_report_excel(start_date: str, end_date: str, group_by: str = "ticket"):
        # 1. Fetch the data using get_monthly_report
        data = ReportsService.get_monthly_report(start_date, end_date, group_by)
        
        # 2. Create openpyxl workbook
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Design headers and styling (all bold_font except title and header_font)
        title_font = Font(name="Segoe UI", size=16, bold=True, color="FF172B4D")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
        bold_font = Font(name="Segoe UI", size=10, bold=True, color="FF172B4D")
        
        header_fill = PatternFill(start_color="FF0052CC", end_color="FF0052CC", fill_type="solid") # Jira Blue
        zebra_fill = PatternFill(start_color="FFFAFBFC", end_color="FFFAFBFC", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='FF808080'),
            right=Side(style='thin', color='FF808080'),
            top=Side(style='thin', color='FF808080'),
            bottom=Side(style='thin', color='FF808080')
        )
        
        # Title Row
        ws.merge_cells("A1:H1")
        for col_idx in range(1, 9):
            ws.cell(row=1, column=col_idx).border = thin_border
        title_cell = ws["A1"]
        title_cell.value = f"Monthly Work Report ({start_date} to {end_date})"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Blank row with borders
        for col_idx in range(1, 9):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = ""
            cell.border = thin_border
        ws.row_dimensions[2].height = 20
        
        # Helper to format logged time from "H.MM" string back to "Xh YYm"
        def format_time_str(h_mm_str):
            if not h_mm_str:
                return "0h 00m"
            parts = h_mm_str.split('.')
            hrs = int(parts[0] or '0')
            mins = int(parts[1] or '0')
            return f"{hrs}h {mins:02d}m"

        # Helper to parse "H.MM" to minutes
        def parse_hmm_to_mins(h_mm_str):
            if not h_mm_str:
                return 0
            parts = h_mm_str.split('.')
            hrs = int(parts[0] or '0')
            mins = int(parts[1] or '0')
            return hrs * 60 + mins
        
        # Calculate Grand Totals
        total_worked_mins = 0
        total_actual_mins = 0
        if group_by == "developer":
            for dev in data:
                for t in dev.get("tickets", []):
                    total_worked_mins += parse_hmm_to_mins(t["working_hours"])
                    total_actual_mins += parse_hmm_to_mins(t["actual_working_hours"])
        else:
            for ticket in data:
                for u in ticket.get("users", []):
                    total_worked_mins += parse_hmm_to_mins(u["working_hours"])
                    total_actual_mins += parse_hmm_to_mins(u["actual_working_hours"])
                    
        grand_total_worked = format_time_str(f"{total_worked_mins // 60}.{total_worked_mins % 60:02d}")
        grand_total_actual = format_time_str(f"{total_actual_mins // 60}.{total_actual_mins % 60:02d}")

        if group_by == "developer":
            # Developer-wise columns:
            headers = [
                "Ticket No", "Ticket Title", "Project Name", "Developer Name",
                "Worked Time", "Total Worked Time", "Actual Time", "Total Actual Time"
            ]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[3].height = 25
            
            row_idx = 4
            for dev in data:
                dev_name = dev["developer_name"]
                dev_worked = format_time_str(dev["total_working_hours"])
                dev_actual = format_time_str(dev["total_actual_hours"])
                
                tickets = dev.get("tickets", [])
                if not tickets:
                    row_data = ["", "", "", dev_name, "", dev_worked, "", dev_actual]
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = val
                        cell.font = bold_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    row_idx += 1
                else:
                    for t_idx, t in enumerate(tickets):
                        t_no = t["ticket_no"]
                        p_name = t["project_name"] or ""
                        t_title = t["title"]
                        t_worked = format_time_str(t["working_hours"])
                        t_actual = format_time_str(t["actual_working_hours"])
                        
                        row_data = [
                            t_no, 
                            t_title, 
                            p_name, 
                            dev_name if t_idx == 0 else "", 
                            t_worked, 
                            dev_worked if t_idx == 0 else "", 
                            t_actual, 
                            dev_actual if t_idx == 0 else ""
                        ]
                        
                        for col_idx, val in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.value = val
                            cell.font = bold_font
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            if row_idx % 2 == 0:
                                cell.fill = zebra_fill
                        row_idx += 1
        
        else:
            # Ticket-wise columns:
            headers = [
                "Ticket No", "Ticket Title", "Project Name", "Developer Name",
                "Worked Time", "Total Worked Time", "Actual Time", "Total Actual Time"
            ]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[3].height = 25
            
            row_idx = 4
            for ticket in data:
                t_no = ticket["ticket_no"]
                p_name = ticket["project_name"] or ""
                t_title = ticket["title"]
                t_worked = format_time_str(ticket["total_working_hours"])
                t_actual = format_time_str(ticket["total_actual_hours"])
                
                users = ticket.get("users", [])
                if not users:
                    row_data = [t_no, t_title, p_name, "", "", t_worked, "", t_actual]
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = val
                        cell.font = bold_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    row_idx += 1
                else:
                    for u_idx, u in enumerate(users):
                        u_name = u["developer_name"]
                        u_worked = format_time_str(u["working_hours"])
                        u_actual = format_time_str(u["actual_working_hours"])
                        
                        row_data = [
                            t_no if u_idx == 0 else "", 
                            t_title if u_idx == 0 else "", 
                            p_name if u_idx == 0 else "", 
                            u_name, 
                            u_worked, 
                            t_worked if u_idx == 0 else "", 
                            u_actual, 
                            t_actual if u_idx == 0 else ""
                        ]
                        
                        for col_idx, val in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.value = val
                            cell.font = bold_font
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            if row_idx % 2 == 0:
                                cell.fill = zebra_fill
                        row_idx += 1

            # Add Blank Row with borders and centered alignment
            blank_row_data = ["", "", "", "", "", "", "", ""]
            for col_idx, val in enumerate(blank_row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = bold_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1

            # Add Grand Total Row for Ticket-wise Report (columns 6 and 8 only, no total in other columns)
            total_row_data = ["Total", "", "", "", "", grand_total_worked, "", grand_total_actual]
            for col_idx, val in enumerate(total_row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = bold_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def export_daily_report_excel(date: str, login_user_id: int):
        # 1. Fetch data
        data = ReportsService.get_daily_report(date, login_user_id)
        
        # 2. Create openpyxl workbook
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Report"
        
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Design styling
        title_font = Font(name="Segoe UI", size=16, bold=True, color="FF172B4D")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
        bold_font = Font(name="Segoe UI", size=10, bold=True, color="FF172B4D")
        
        header_fill = PatternFill(start_color="FF0052CC", end_color="FF0052CC", fill_type="solid") # Jira Blue
        zebra_fill = PatternFill(start_color="FFFAFBFC", end_color="FFFAFBFC", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='FF808080'),
            right=Side(style='thin', color='FF808080'),
            top=Side(style='thin', color='FF808080'),
            bottom=Side(style='thin', color='FF808080')
        )
        
        # Title Row (A1:G1)
        ws.merge_cells("A1:G1")
        for col_idx in range(1, 8):
            ws.cell(row=1, column=col_idx).border = thin_border
        title_cell = ws["A1"]
        title_cell.value = f"Daily Work Report ({date})"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Blank row with borders (A2:G2)
        for col_idx in range(1, 8):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = ""
            cell.border = thin_border
        ws.row_dimensions[2].height = 20
        
        # Helper to format logged time from "H.MM" string back to "Xh YYm"
        def format_time_str(h_mm_str):
            if not h_mm_str:
                return "0h 00m"
            parts = h_mm_str.split('.')
            hrs = int(parts[0] or '0')
            mins = int(parts[1] or '0')
            return f"{hrs}h {mins:02d}m"

        # Helper to parse "H.MM" to minutes
        def parse_hmm_to_mins(h_mm_str):
            if not h_mm_str:
                return 0
            parts = h_mm_str.split('.')
            hrs = int(parts[0] or '0')
            mins = int(parts[1] or '0')
            return hrs * 60 + mins

        # Headers Row 3
        headers = ["Developer Name", "Target Hours", "Ticket No", "Project Name", "Ticket Title", "Worked Time", "Actual Time"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[3].height = 25
        
        # Data Rows
        row_idx = 4
        total_worked_mins = 0
        total_actual_mins = 0
        
        for dev in data:
            dev_name = dev["developer_name"]
            target_hours_val = float(dev["total_working_hours"] or 0)
            target_hours_display = f"{target_hours_val}h"
            
            tickets = dev.get("tickets", [])
            if not tickets:
                row_data = [dev_name, target_hours_display, "", "", "", "", ""]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = val
                    cell.font = bold_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                row_idx += 1
            else:
                for t_idx, t in enumerate(tickets):
                    t_no = t["ticket_no"]
                    p_name = t["project_name"] or ""
                    t_title = t["title"]
                    t_worked = format_time_str(t["today_working_hours"])
                    t_actual = format_time_str(t["actual_working_hours"])
                    
                    total_worked_mins += parse_hmm_to_mins(t["today_working_hours"])
                    total_actual_mins += parse_hmm_to_mins(t["actual_working_hours"])
                    
                    row_data = [
                        dev_name if t_idx == 0 else "", 
                        target_hours_display if t_idx == 0 else "", 
                        t_no, p_name, t_title, t_worked, t_actual
                    ]
                    
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = val
                        cell.font = bold_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        if row_idx % 2 == 0:
                            cell.fill = zebra_fill
                    row_idx += 1
                    
        # Add Blank Row with borders
        blank_row_data = ["", "", "", "", "", "", ""]
        for col_idx, val in enumerate(blank_row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1
        
        # Add Grand Total Row (columns 6 and 7 only)
        grand_total_worked = format_time_str(f"{total_worked_mins // 60}.{total_worked_mins % 60:02d}")
        grand_total_actual = format_time_str(f"{total_actual_mins // 60}.{total_actual_mins % 60:02d}")
        
        total_row_data = ["Total", "", "", "", "", grand_total_worked, grand_total_actual]
        for col_idx, val in enumerate(total_row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream
